import io
import os
import re
import shutil
import tempfile
import uuid
from datetime import datetime, timezone
from flask import Blueprint, request, jsonify, send_file, current_app
from werkzeug.utils import secure_filename
from app import db
from app.models import FitxaTecnica, VersioFitxa, Distribucio, DestiDistribucio
from app.auth import login_required, rol_requerit

fitxes_bp = Blueprint('fitxes', __name__)


def _validar_art_codi(art_codi):
    """Valida que art_codi només conté caràcters segurs (lletres, números, guions)."""
    if not art_codi:
        return False
    if not re.match(r'^[A-Za-z0-9_\-\.]+$', art_codi):
        return False
    if '..' in art_codi:
        return False
    return True

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', 'uploads')
TEMP_IMG_DIR = os.path.join(UPLOAD_DIR, '_temp_imatges')

ALLOWED_IMG_EXT = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg'}


def _ensure_upload_dir(art_codi, num_versio):
    path = os.path.join(UPLOAD_DIR, art_codi, f'v{num_versio}')
    os.makedirs(path, exist_ok=True)
    return path


def _parse_data(val):
    """Parseja un valor de data del frontend o del Word.
    Accepta ISO ('2026-06-04', '2026-06-04T...') o format 'dd/mm/aaaa'.
    Retorna datetime o None."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    # ISO
    try:
        return datetime.fromisoformat(s.replace('Z', '+00:00'))
    except ValueError:
        pass
    # dd/mm/aaaa
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_num_versio(val, default=1):
    """Converteix 'rev' del Word a enter. Default si no és vàlid."""
    if val is None or val == '':
        return default
    try:
        n = int(str(val).strip())
        return n if n >= 0 else default
    except (ValueError, TypeError):
        return default


def _esborrar_destins(fitxa, dest_ids):
    """Esborra el PDF d'una fitxa dels destins indicats (FTP/xarxa).
    Reutilitza la informació de Distribucio.missatge_error per a obtenir
    els noms reals de fitxer distribuïts. Retorna llista de resultats."""
    if not dest_ids:
        return []

    resultats = []
    for desti_id in dest_ids:
        desti = DestiDistribucio.query.get(desti_id)
        if not desti:
            continue
        config = desti.configuracio or {}

        dist_ok = Distribucio.query.join(VersioFitxa).filter(
            VersioFitxa.fitxa_id == fitxa.id,
            Distribucio.desti_id == desti_id,
            Distribucio.estat == 'ok',
        ).all()
        filenames = set()
        for d in dist_ok:
            if d.missatge_error:
                name = d.missatge_error.split('/')[-1].split('\\')[-1]
                if name:
                    filenames.add(name)
        if not filenames:
            filenames.add(f'{fitxa.art_codi}.pdf')

        for fname in filenames:
            if desti.tipus == 'ftp':
                from app.services.ftp_distributor import eliminar_ftp
                result = eliminar_ftp(fitxa.art_codi, config, fname)
            elif desti.tipus == 'xarxa':
                from app.services.smb_distributor import eliminar_xarxa
                result = eliminar_xarxa(fitxa.art_codi, config, fname)
            else:
                result = {'ok': False, 'error': f"Tipus {desti.tipus} no suportat"}
            resultats.append({'desti': desti.nom, 'fitxer': fname, **result})

    return resultats


def _guardar_imatges_temp(imatges):
    """Guarda imatges extretes del Word a un directori temporal.
    Retorna el token i la llista de filenames finals."""
    if not imatges:
        return None, []
    token = uuid.uuid4().hex
    temp_dir = os.path.join(TEMP_IMG_DIR, token)
    os.makedirs(temp_dir, exist_ok=True)

    saved = []
    for idx, img in enumerate(imatges):
        filename = secure_filename(img.get('filename') or f'image_{idx + 1}.png')
        ext = os.path.splitext(filename)[1].lower()
        if ext not in ALLOWED_IMG_EXT:
            continue
        dest = os.path.join(temp_dir, filename)
        # Evitar col·lisions
        if os.path.exists(dest):
            name, extension = os.path.splitext(filename)
            filename = f'{name}_{idx + 1}{extension}'
            dest = os.path.join(temp_dir, filename)
        with open(dest, 'wb') as f:
            f.write(img['blob'])
        saved.append(filename)

    return token, saved


def _moure_imatges_temp(token, art_codi):
    """Mou imatges del directori temporal a uploads/<art_codi>/img/.
    Retorna llista d'URLs finals."""
    if not token:
        return []
    temp_dir = os.path.join(TEMP_IMG_DIR, token)
    if not os.path.isdir(temp_dir):
        return []

    img_dir = os.path.join(UPLOAD_DIR, art_codi, 'img')
    os.makedirs(img_dir, exist_ok=True)

    urls = []
    for filename in os.listdir(temp_dir):
        ext = os.path.splitext(filename)[1].lower()
        if ext not in ALLOWED_IMG_EXT:
            continue
        src = os.path.join(temp_dir, filename)
        dest_name = filename
        dest = os.path.join(img_dir, dest_name)
        # Evitar sobreescriure
        if os.path.exists(dest):
            name, extension = os.path.splitext(dest_name)
            dest_name = f'{name}_{int(datetime.now(timezone.utc).timestamp())}{extension}'
            dest = os.path.join(img_dir, dest_name)
        try:
            shutil.move(src, dest)
        except OSError:
            continue
        urls.append(dest_name)

    # Netejar directori temp
    try:
        shutil.rmtree(temp_dir, ignore_errors=True)
    except OSError:
        pass

    return urls


SORT_COLS = {
    'art_codi': FitxaTecnica.art_codi,
    'nom_producte': FitxaTecnica.nom_producte,
    'categoria': FitxaTecnica.categoria,
    'estat': FitxaTecnica.estat,
    'updated_at': FitxaTecnica.updated_at,
}


@fitxes_bp.route('/fitxes', methods=['GET'])
@login_required
def llistar_fitxes():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    cerca = request.args.get('cerca', '', type=str)

    query = FitxaTecnica.query

    estat = request.args.get('estat', '', type=str)
    if estat:
        query = query.filter(FitxaTecnica.estat == estat)

    if cerca:
        filtre = f'%{cerca}%'
        query = query.filter(
            db.or_(
                FitxaTecnica.art_codi.ilike(filtre),
                FitxaTecnica.nom_producte.ilike(filtre),
            )
        )

    sort_by = request.args.get('sort_by', 'updated_at', type=str)
    sort_order = request.args.get('sort_order', 'desc', type=str).lower()
    direction = 'asc' if sort_order == 'asc' else 'desc'

    if sort_by == 'versio_activa':
        # Subquery: num_versio de la versió activa per cada fitxa
        sub = db.session.query(
            VersioFitxa.fitxa_id.label('fid'),
            VersioFitxa.num_versio.label('vnum'),
        ).filter(VersioFitxa.activa == True).subquery()
        query = query.outerjoin(sub, sub.c.fid == FitxaTecnica.id)
        col = sub.c.vnum
        query = query.order_by(col.asc() if direction == 'asc' else col.desc(),
                               FitxaTecnica.id.asc())
    else:
        col = SORT_COLS.get(sort_by, FitxaTecnica.updated_at)
        query = query.order_by(col.asc() if direction == 'asc' else col.desc(),
                               FitxaTecnica.id.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # Afegir resum de distribució i versió activa per cada fitxa
    fitxes_data = []
    for f in pagination.items:
        d = f.to_dict()
        # Versió activa
        va = VersioFitxa.query.filter_by(fitxa_id=f.id, activa=True).first()
        d['versio_activa'] = va.num_versio if va else None
        # Resum distribucions de la versió activa
        if va:
            dists = Distribucio.query.filter_by(versio_id=va.id).all()
            total_destins = DestiDistribucio.query.filter_by(actiu=True).count()
            ok_count = sum(1 for dd in dists if dd.estat == 'ok')
            error_count = sum(1 for dd in dists if dd.estat == 'error')
            d['dist_resum'] = {
                'total_destins': total_destins,
                'ok': ok_count,
                'error': error_count,
                'pendent': total_destins - ok_count - error_count,
            }
        else:
            d['dist_resum'] = None
        fitxes_data.append(d)

    return jsonify({
        'fitxes': fitxes_data,
        'total': pagination.total,
        'pages': pagination.pages,
        'page': page,
    })


@fitxes_bp.route('/fitxes/<int:fitxa_id>', methods=['GET'])
@login_required
def detall_fitxa(fitxa_id):
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    return jsonify(fitxa.to_dict(include_versions=True))


@fitxes_bp.route('/fitxes', methods=['POST'])
@rol_requerit('admin', 'editor')
def crear_fitxa():
    data = request.get_json()

    if not data or not data.get('art_codi') or not data.get('nom_producte'):
        return jsonify({'error': "Camps obligatoris: art_codi, nom_producte"}), 400

    if not _validar_art_codi(data['art_codi']):
        return jsonify({'error': "Codi d'article invàlid. Només lletres, números, guions i guions baixos."}), 400

    if FitxaTecnica.query.filter_by(art_codi=data['art_codi']).first():
        return jsonify({'error': f"Ja existeix una fitxa amb codi {data['art_codi']}"}), 409

    fitxa = FitxaTecnica(
        art_codi=data['art_codi'],
        nom_producte=data['nom_producte'],
        categoria=data.get('categoria', ''),
        estat='publicada',
        created_by=request.usuari.get('email', ''),
    )
    db.session.add(fitxa)
    db.session.flush()

    # Parsejar dates de capçalera si venen del Word
    data_rev_dt = _parse_data(data.get('data_revisio'))
    data_comp_dt = _parse_data(data.get('data_comprovacio'))
    num_versio = _parse_num_versio(data.get('rev'), default=1)

    versio = VersioFitxa(
        fitxa_id=fitxa.id,
        num_versio=num_versio,
        descripcio_canvi=data.get('descripcio_canvi', 'Creació inicial'),
        contingut=data.get('contingut', {}),
        data_revisio=data_rev_dt,
        data_comprovacio=data_comp_dt,
        created_by=request.usuari.get('email', ''),
        activa=True,
        estat_versio='publicada',
    )
    db.session.add(versio)
    db.session.commit()

    # Moure imatges del Word (si hi ha temp_token) a uploads/<art_codi>/img/
    temp_token = data.get('imatges_temp_token')
    if temp_token:
        _moure_imatges_temp(temp_token, fitxa.art_codi)

    return jsonify(fitxa.to_dict(include_versions=True)), 201


@fitxes_bp.route('/fitxes/upload-word', methods=['POST'])
@rol_requerit('admin', 'editor')
def upload_word():
    """Puja un .docx, extreu les dades i crea/actualitza la fitxa."""
    if 'file' not in request.files:
        return jsonify({'error': "Cal enviar un fitxer .docx"}), 400

    file = request.files['file']
    if not file.filename.endswith('.docx'):
        return jsonify({'error': "El fitxer ha de ser .docx"}), 400

    # Guardar temporalment per parsejar
    with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    try:
        from app.services.word_parser import parse_docx
        result = parse_docx(tmp_path)
    except Exception as e:
        os.unlink(tmp_path)
        return jsonify({'error': f"Error parsejant el Word: {str(e)}"}), 400

    art_codi = result['art_codi']
    if not art_codi:
        os.unlink(tmp_path)
        return jsonify({'error': "No s'ha trobat el codi de referència al document"}), 400

    # Guardar imatges incrustades a directori temporal
    imatges_token, imatges_noms = _guardar_imatges_temp(result.get('imatges', []))

    # Comprovar si la fitxa ja existeix
    fitxa_existent = FitxaTecnica.query.filter_by(art_codi=art_codi).first()

    if fitxa_existent:
        # Retornar dades extretes + info de la fitxa existent per confirmar
        os.unlink(tmp_path)
        return jsonify({
            'existent': True,
            'fitxa': fitxa_existent.to_dict(include_versions=True),
            'dades_extretes': result['contingut'],
            'rev': result['rev'],
            'data_revisio': result['data_revisio'],
            'data_comprovacio': result['data_comprovacio'],
            'imatges_temp_token': imatges_token,
            'imatges_temp_noms': imatges_noms,
            'message': f"La fitxa {art_codi} ja existeix. Vols crear una nova versió?",
        }), 200

    # Fitxa nova: retornar dades extretes perquè l'usuari confirmi
    os.unlink(tmp_path)
    return jsonify({
        'existent': False,
        'dades_extretes': result['contingut'],
        'rev': result['rev'],
        'data_revisio': result['data_revisio'],
        'data_comprovacio': result['data_comprovacio'],
        'art_codi': art_codi,
        'nom_producte': result['contingut'].get('denominacio_comercial', ''),
        'imatges_temp_token': imatges_token,
        'imatges_temp_noms': imatges_noms,
    }), 200


@fitxes_bp.route('/fitxes/<int:fitxa_id>/upload-docx', methods=['POST'])
@rol_requerit('admin', 'editor')
def guardar_docx_versio(fitxa_id):
    """Guarda el fitxer .docx original associat a una versió."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)

    if 'file' not in request.files:
        return jsonify({'error': "Cal enviar un fitxer .docx"}), 400

    file = request.files['file']
    vid = request.form.get('versio_id')
    if not vid:
        return jsonify({'error': "Cal indicar versio_id"}), 400

    versio = VersioFitxa.query.filter_by(fitxa_id=fitxa_id, id=int(vid)).first_or_404()

    upload_path = _ensure_upload_dir(fitxa.art_codi, versio.num_versio)
    filename = secure_filename(file.filename) or f'{fitxa.art_codi}.docx'
    filepath = os.path.join(upload_path, filename)
    file.save(filepath)

    versio.fitxer_docx = filepath
    db.session.commit()

    return jsonify({'message': 'Fitxer guardat', 'path': filepath})


@fitxes_bp.route('/fitxes/<int:fitxa_id>', methods=['PUT'])
@rol_requerit('admin', 'editor')
def editar_fitxa(fitxa_id):
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    data = request.get_json()

    if 'nom_producte' in data:
        fitxa.nom_producte = data['nom_producte']
    if 'categoria' in data:
        fitxa.categoria = data['categoria']

    db.session.commit()
    return jsonify(fitxa.to_dict())


@fitxes_bp.route('/fitxes/<int:fitxa_id>/duplicar', methods=['POST'])
@rol_requerit('admin', 'editor')
def duplicar_fitxa(fitxa_id):
    """Duplica una fitxa amb un nou codi d'article. Només copia contingut, no versions ni distribucions."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    data = request.get_json() or {}

    nou_codi = (data.get('art_codi') or '').strip()
    if not nou_codi:
        return jsonify({'error': "Cal indicar el nou codi d'article"}), 400

    if not _validar_art_codi(nou_codi):
        return jsonify({'error': "Codi d'article invàlid. Només lletres, números, guions i guions baixos."}), 400

    if FitxaTecnica.query.filter_by(art_codi=nou_codi).first():
        return jsonify({'error': f"Ja existeix una fitxa amb codi {nou_codi}"}), 409

    # Obtenir contingut de la versió activa
    versio_activa = VersioFitxa.query.filter_by(fitxa_id=fitxa_id, activa=True).first()
    contingut = dict(versio_activa.contingut) if versio_activa and versio_activa.contingut else {}
    contingut['codi_referencia'] = nou_codi

    nova_fitxa = FitxaTecnica(
        art_codi=nou_codi,
        nom_producte=data.get('nom_producte', fitxa.nom_producte),
        categoria=fitxa.categoria,
        estat='publicada',
        es_client=fitxa.es_client,
        observacions=fitxa.observacions or '',
        created_by=request.usuari.get('email', ''),
    )
    db.session.add(nova_fitxa)
    db.session.flush()

    nova_versio = VersioFitxa(
        fitxa_id=nova_fitxa.id,
        num_versio=1,
        descripcio_canvi=f"Duplicada de {fitxa.art_codi}",
        contingut=contingut,
        created_by=request.usuari.get('email', ''),
        activa=True,
        estat_versio='publicada',
    )
    db.session.add(nova_versio)
    db.session.commit()

    return jsonify(nova_fitxa.to_dict(include_versions=True)), 201


@fitxes_bp.route('/fitxes/<int:fitxa_id>/observacions', methods=['PUT'])
@rol_requerit('admin', 'editor')
def actualitzar_observacions(fitxa_id):
    """Actualitza les observacions d'una fitxa."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    data = request.get_json() or {}
    fitxa.observacions = data.get('observacions', '')
    db.session.commit()
    return jsonify({'message': 'Observacions actualitzades', 'observacions': fitxa.observacions})


ESTATS_VALIDS = {'esborrany', 'publicada', 'obsoleta', 'inactiva'}


@fitxes_bp.route('/fitxes/<int:fitxa_id>/estat', methods=['PATCH'])
@rol_requerit('admin', 'editor')
def canviar_estat(fitxa_id):
    """Canvia l'estat d'una fitxa. Si passa a 'inactiva' i s'indiquen
    destins, esborra el PDF d'aquests destins i ho registra a
    RegistreEliminacio per a audit trail."""
    from app.models import RegistreEliminacio, Usuari

    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    data = request.get_json() or {}

    nou_estat = (data.get('estat') or '').strip().lower()
    if nou_estat not in ESTATS_VALIDS:
        return jsonify({
            'error': f"Estat invàlid. Vàlids: {', '.join(sorted(ESTATS_VALIDS))}"
        }), 400

    # Editor només pot fer transicions no destructives
    rol = (request.usuari or {}).get('rol', '')
    if rol != 'admin' and nou_estat in ('inactiva', 'obsoleta'):
        return jsonify({
            'error': "Només admin pot marcar fitxes com a inactiva o obsoleta"
        }), 403

    esborrar_destins = data.get('esborrar_destins', []) or []
    motiu = (data.get('motiu') or '').strip()

    dest_resultats = []
    if nou_estat == 'inactiva' and esborrar_destins:
        dest_resultats = _esborrar_destins(fitxa, esborrar_destins)

        # Audit trail: registrar a RegistreEliminacio
        usuari = Usuari.query.filter_by(email=request.usuari.get('email')).first()
        versions_list = fitxa.versions.all()
        registre = RegistreEliminacio(
            art_codi=fitxa.art_codi,
            nom_producte=fitxa.nom_producte,
            num_versions=len(versions_list),
            ultima_versio=max((v.num_versio for v in versions_list), default=0),
            motiu=f"Inactivada (canvi d'estat): {motiu}" if motiu else "Inactivada (canvi d'estat)",
            esborrat_ftp=True,
            eliminat_per=usuari.nom if usuari else (request.usuari.get('email') or ''),
        )
        db.session.add(registre)

    fitxa.estat = nou_estat
    db.session.commit()

    return jsonify({
        'fitxa': fitxa.to_dict(),
        'destins': dest_resultats,
    })


@fitxes_bp.route('/fitxes/<int:fitxa_id>', methods=['DELETE'])
@rol_requerit('admin')
def eliminar_fitxa(fitxa_id):
    from app.models import Distribucio, Usuari, RegistreEliminacio
    import shutil

    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    data = request.get_json() or {}

    # Validar motiu obligatori
    motiu = (data.get('motiu') or '').strip()
    if not motiu:
        return jsonify({'error': "Cal indicar un motiu per eliminar la fitxa"}), 400

    # Validar contrasenya de l'usuari
    password = data.get('password', '')
    if not password:
        return jsonify({'error': "Cal confirmar amb la teva contrasenya"}), 400

    usuari = Usuari.query.filter_by(email=request.usuari.get('email')).first()
    if not usuari or not usuari.check_password(password):
        return jsonify({'error': "Contrasenya incorrecta"}), 403

    esborrar_destins = data.get('esborrar_destins', [])
    nomes_inactivar = data.get('nomes_inactivar', False)

    # Info per al registre
    versions_list = fitxa.versions.all()
    num_versions = len(versions_list)
    ultima_versio = max((v.num_versio for v in versions_list), default=0)

    # Esborrar dels destins seleccionats
    dest_resultats = _esborrar_destins(fitxa, esborrar_destins)

    # Registrar l'acció
    accio_text = 'Inactivada' if nomes_inactivar else 'Eliminada'
    registre = RegistreEliminacio(
        art_codi=fitxa.art_codi,
        nom_producte=fitxa.nom_producte,
        num_versions=num_versions,
        ultima_versio=ultima_versio,
        motiu=f"{accio_text}: {motiu}",
        esborrat_ftp=len(esborrar_destins) > 0,
        eliminat_per=usuari.nom,
    )
    db.session.add(registre)

    if nomes_inactivar:
        # Només canviar l'estat a 'inactiva', no eliminar res
        fitxa.estat = 'inactiva'
        db.session.commit()
        return jsonify({
            'message': f'Fitxa {fitxa.art_codi} marcada com a inactiva',
            'destins': dest_resultats,
        }), 200
    else:
        # Eliminar distribucions, versions i fitxa
        for versio in versions_list:
            Distribucio.query.filter_by(versio_id=versio.id).delete()
        VersioFitxa.query.filter_by(fitxa_id=fitxa_id).delete()
        db.session.delete(fitxa)
        db.session.commit()

        # Eliminar fitxers locals (uploads)
        upload_path = os.path.join(UPLOAD_DIR, fitxa.art_codi)
        if os.path.exists(upload_path):
            shutil.rmtree(upload_path, ignore_errors=True)

        return jsonify({
            'message': 'Fitxa eliminada',
            'destins': dest_resultats,
        }), 200


@fitxes_bp.route('/fitxes/eliminacions', methods=['GET'])
@rol_requerit('admin')
def llistar_eliminacions():
    """Historial de fitxes eliminades."""
    from app.models import RegistreEliminacio
    registres = RegistreEliminacio.query.order_by(
        RegistreEliminacio.eliminat_at.desc()
    ).all()
    return jsonify([r.to_dict() for r in registres])


@fitxes_bp.route('/fitxes/<int:fitxa_id>/imatges', methods=['POST'])
@rol_requerit('admin', 'editor')
def pujar_imatge(fitxa_id):
    """Puja una imatge associada a una fitxa (certificacions, segells, etc.)."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)

    if 'file' not in request.files:
        return jsonify({'error': "Cal enviar un fitxer d'imatge"}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': "Fitxer buit"}), 400

    # Validar extensió
    allowed = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg'}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        return jsonify({'error': f"Format no permès. Formats vàlids: {', '.join(allowed)}"}), 400

    # Guardar a uploads/{art_codi}/img/
    img_dir = os.path.join(UPLOAD_DIR, fitxa.art_codi, 'img')
    os.makedirs(img_dir, exist_ok=True)

    filename = secure_filename(file.filename)
    # Evitar sobreescriure: afegir timestamp si ja existeix
    filepath = os.path.join(img_dir, filename)
    if os.path.exists(filepath):
        name, extension = os.path.splitext(filename)
        filename = f"{name}_{int(datetime.now(timezone.utc).timestamp())}{extension}"
        filepath = os.path.join(img_dir, filename)

    file.save(filepath)

    url = f'/api/fitxes/{fitxa_id}/imatges/{filename}'
    return jsonify({'url': url, 'filename': filename}), 201


@fitxes_bp.route('/fitxes/<int:fitxa_id>/imatges/from-temp', methods=['POST'])
@rol_requerit('admin', 'editor')
def imatges_from_temp(fitxa_id):
    """Mou imatges extretes prèviament del Word (temp_token) a uploads/<art_codi>/img/."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    data = request.get_json() or {}
    token = (data.get('temp_token') or '').strip()
    if not token:
        return jsonify({'error': 'Cal indicar temp_token'}), 400

    # Validar token (només hex generat per uuid)
    if not re.match(r'^[a-f0-9]{32}$', token):
        return jsonify({'error': 'Token invàlid'}), 400

    nous = _moure_imatges_temp(token, fitxa.art_codi)
    urls = [
        {'filename': n, 'url': f'/api/fitxes/{fitxa_id}/imatges/{n}'} for n in nous
    ]
    return jsonify({'imatges': urls, 'total': len(urls)}), 200


@fitxes_bp.route('/fitxes/<int:fitxa_id>/imatges/<path:filename>', methods=['GET'])
def servir_imatge(fitxa_id, filename):
    """Serveix una imatge associada a una fitxa."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    img_dir = os.path.join(UPLOAD_DIR, fitxa.art_codi, 'img')
    filepath = os.path.join(img_dir, secure_filename(filename))

    if not os.path.exists(filepath):
        return jsonify({'error': 'Imatge no trobada'}), 404

    return send_file(filepath)


@fitxes_bp.route('/fitxes/<int:fitxa_id>/imatges', methods=['GET'])
@login_required
def llistar_imatges(fitxa_id):
    """Llista les imatges d'una fitxa."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)
    img_dir = os.path.join(UPLOAD_DIR, fitxa.art_codi, 'img')

    if not os.path.exists(img_dir):
        return jsonify([])

    imatges = []
    for f in os.listdir(img_dir):
        ext = os.path.splitext(f)[1].lower()
        if ext in {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg'}:
            imatges.append({
                'filename': f,
                'url': f'/api/fitxes/{fitxa_id}/imatges/{f}',
            })

    return jsonify(imatges)


@fitxes_bp.route('/fitxes/<int:fitxa_id>/pdf', methods=['GET'])
@login_required
def descarregar_pdf(fitxa_id):
    """Genera el PDF des de les dades de l'app (contingut JSON).
    Paràmetre ?original=1 retorna el PDF original del FTP si existeix."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)

    vid = request.args.get('versio_id', None, type=int)
    if vid:
        versio = VersioFitxa.query.filter_by(fitxa_id=fitxa_id, id=vid).first_or_404()
    else:
        versio = VersioFitxa.query.filter_by(fitxa_id=fitxa_id, activa=True).first()
        if not versio:
            versio = VersioFitxa.query.filter_by(fitxa_id=fitxa_id)\
                .order_by(VersioFitxa.num_versio.desc()).first()

    if not versio:
        return jsonify({'error': "No hi ha versions"}), 404

    # Si demanen l'original del FTP
    if request.args.get('original') == '1':
        if versio.fitxer_pdf and os.path.exists(versio.fitxer_pdf):
            return send_file(
                versio.fitxer_pdf,
                mimetype='application/pdf',
                as_attachment=False,
                download_name=f'{fitxa.art_codi}_original.pdf',
            )
        return jsonify({'error': "No hi ha PDF original"}), 404

    # Comprovar cache: si existeix PDF generat per aquesta versió
    cache_dir = os.path.join(UPLOAD_DIR, fitxa.art_codi, f'v{versio.num_versio}')
    cache_path = os.path.join(cache_dir, f'{fitxa.art_codi}_generat.pdf')
    force_regen = request.args.get('regen', '0') == '1'

    if force_regen and os.path.exists(cache_path):
        os.remove(cache_path)

    if not os.path.exists(cache_path):
        # Generar PDF des de les dades de l'app
        from app.services.pdf_generator import generar_pdf

        contingut = versio.contingut or {}
        if 'codi_referencia' not in contingut:
            contingut['codi_referencia'] = fitxa.art_codi
        if 'denominacio_comercial' not in contingut:
            contingut['denominacio_comercial'] = fitxa.nom_producte

        data_rev = ''
        data_comp = ''
        if versio.data_comprovacio:
            data_comp = versio.data_comprovacio.strftime('%d/%m/%Y')
        if versio.data_revisio:
            data_rev = versio.data_revisio.strftime('%d/%m/%Y')
        elif versio.created_at:
            data_rev = versio.created_at.strftime('%d/%m/%Y')

        pdf_bytes = generar_pdf(contingut, versio.num_versio, data_rev, data_comp)

        # Guardar a cache
        os.makedirs(cache_dir, exist_ok=True)
        with open(cache_path, 'wb') as f:
            f.write(pdf_bytes)

    return send_file(
        cache_path,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=f'{fitxa.art_codi}.pdf',
    )


@fitxes_bp.route('/fitxes/<int:fitxa_id>/verificar', methods=['GET'])
@login_required
def verificar_fitxa(fitxa_id):
    """Compara les dades de l'app amb el PDF original del FTP."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)

    versio = VersioFitxa.query.filter_by(fitxa_id=fitxa_id, activa=True).first()
    if not versio:
        versio = VersioFitxa.query.filter_by(fitxa_id=fitxa_id)\
            .order_by(VersioFitxa.num_versio.desc()).first()

    if not versio:
        return jsonify({'error': "No hi ha versions"}), 404

    if not versio.fitxer_pdf or not os.path.exists(versio.fitxer_pdf):
        return jsonify({'error': "No hi ha PDF original per comparar"}), 404

    from app.services.pdf_parser import parse_pdf

    try:
        pdf_data = parse_pdf(versio.fitxer_pdf)
    except Exception as e:
        return jsonify({'error': f"Error parsejant el PDF: {str(e)}"}), 500

    pdf_contingut = pdf_data['contingut']
    app_contingut = versio.contingut or {}

    diferencies = []

    # Comparar camps de text
    all_keys = set(list(pdf_contingut.keys()) + list(app_contingut.keys()))
    for key in sorted(all_keys):
        pdf_val = pdf_contingut.get(key)
        app_val = app_contingut.get(key)

        if isinstance(pdf_val, list) or isinstance(app_val, list):
            # Comparar taules
            pdf_rows = pdf_val if isinstance(pdf_val, list) else []
            app_rows = app_val if isinstance(app_val, list) else []

            if len(pdf_rows) != len(app_rows):
                diferencies.append({
                    'camp': key,
                    'tipus': 'taula',
                    'detall': f"Nombre de files diferent: PDF={len(pdf_rows)}, App={len(app_rows)}",
                    'pdf': pdf_rows,
                    'app': app_rows,
                })
            else:
                for i, (pr, ar) in enumerate(zip(pdf_rows, app_rows)):
                    p_param = pr.get('parametre', '')
                    a_param = ar.get('parametre', '')
                    p_val = pr.get('valor', '')
                    a_val = ar.get('valor', '')
                    if p_param != a_param or p_val != a_val:
                        diferencies.append({
                            'camp': key,
                            'tipus': 'taula_fila',
                            'fila': i + 1,
                            'detall': f"Fila {i+1} diferent",
                            'pdf': f"{p_param}: {p_val}",
                            'app': f"{a_param}: {a_val}",
                        })
        else:
            pdf_str = str(pdf_val or '').strip()
            app_str = str(app_val or '').strip()

            if not pdf_str and not app_str:
                continue

            if pdf_str != app_str:
                diferencies.append({
                    'camp': key,
                    'tipus': 'text',
                    'pdf': pdf_str[:200] if pdf_str else '(buit)',
                    'app': app_str[:200] if app_str else '(buit)',
                })

    return jsonify({
        'art_codi': fitxa.art_codi,
        'pdf_rev': pdf_data.get('rev', ''),
        'app_rev': versio.num_versio,
        'total_diferencies': len(diferencies),
        'diferencies': diferencies,
        'pdf_camps': len(pdf_contingut),
        'app_camps': len(app_contingut),
    })


@fitxes_bp.route('/fitxes/<int:fitxa_id>/docx', methods=['GET'])
@login_required
def descarregar_docx(fitxa_id):
    """Descarrega el .docx original d'una versió."""
    fitxa = db.get_or_404(FitxaTecnica, fitxa_id)

    vid = request.args.get('versio_id', None, type=int)
    if vid:
        versio = VersioFitxa.query.filter_by(fitxa_id=fitxa_id, id=vid).first_or_404()
    else:
        versio = VersioFitxa.query.filter_by(fitxa_id=fitxa_id, activa=True).first()

    if not versio or not versio.fitxer_docx:
        return jsonify({'error': "No hi ha fitxer .docx per aquesta versió"}), 404

    if not os.path.exists(versio.fitxer_docx):
        return jsonify({'error': "Fitxer no trobat al servidor"}), 404

    return send_file(
        versio.fitxer_docx,
        as_attachment=True,
        download_name=f'{fitxa.art_codi}_v{versio.num_versio}.docx',
    )


def _build_control_revisions():
    """Construeix les dades de control de revisions per a fitxes que tenen PDF (del FTP).
    Optimitzat amb eager loading per evitar N+1 queries."""
    # Només fitxes que tenen almenys una versió amb PDF
    fitxa_ids_amb_pdf = db.session.query(VersioFitxa.fitxa_id).filter(
        VersioFitxa.fitxer_pdf != None, VersioFitxa.fitxer_pdf != ''
    ).distinct().subquery()

    fitxes = FitxaTecnica.query.filter(
        FitxaTecnica.id.in_(fitxa_ids_amb_pdf)
    ).order_by(FitxaTecnica.art_codi).all()

    resultats = []

    for fitxa in fitxes:
        versions_list = sorted(fitxa.versions.all(),
                               key=lambda v: v.num_versio)

        activa = next((v for v in versions_list if v.activa), None)
        ultima = versions_list[-1] if versions_list else None
        versio_ref = activa or ultima

        contingut = versio_ref.contingut if versio_ref else {}
        if not contingut:
            contingut = {}

        def trobar_param(taula_nom, param_prefix):
            taula = contingut.get(taula_nom, [])
            if not isinstance(taula, list):
                return ''
            for row in taula:
                if isinstance(row, dict) and param_prefix.lower() in row.get('parametre', '').lower():
                    return row.get('valor', '')
            return ''

        historial = [{
            'num_versio': v.num_versio,
            'descripcio': v.descripcio_canvi,
            'data': v.created_at.strftime('%d/%m/%Y') if v.created_at else '',
            'autor': v.created_by or '',
        } for v in versions_list]

        resultats.append({
            'id': fitxa.id,
            'art_codi': fitxa.art_codi,
            'nom_producte': fitxa.nom_producte,
            'categoria': fitxa.categoria or '',
            'estat': fitxa.estat,
            'es_client': fitxa.es_client or False,
            'observacions': fitxa.observacions or '',
            'revisio': versio_ref.num_versio if versio_ref else 0,
            'data_revisio': (
                versio_ref.data_revisio.strftime('%d/%m/%Y')
                if versio_ref and versio_ref.data_revisio
                else (versio_ref.created_at.strftime('%d/%m/%Y') if versio_ref and versio_ref.created_at else '')
            ),
            'data_comprovacio': versio_ref.data_comprovacio.strftime('%d/%m/%Y') if versio_ref and versio_ref.data_comprovacio else '',
            'denominacio_juridica': _clean_html(contingut.get('denominacio_juridica', '')),
            'composicio': _clean_html(contingut.get('ingredients', '') or contingut.get('composicio', '')),
            'vida_util': _clean_html(contingut.get('vida_util', '')),
            'w': trobar_param('reologiques', 'W'),
            'pl': trobar_param('reologiques', 'P/L'),
            'proteina': trobar_param('fisicoquimiques', 'Prote'),
            'gluten': trobar_param('fisicoquimiques', 'Gluten'),
            'cendres': trobar_param('fisicoquimiques', 'Ceniz') or trobar_param('fisicoquimiques', 'Cendr'),
            'estat_versio': versio_ref.estat_versio if versio_ref else '',
            'ultima_accio': versio_ref.created_at.isoformat() if versio_ref and versio_ref.created_at else '',
            'ultima_accio_per': versio_ref.created_by if versio_ref else '',
            'historial': historial,
        })

    return resultats


def _clean_html(text):
    """Neteja tags HTML d'un text."""
    if not text:
        return ''
    return re.sub(r'<[^>]+>', '', str(text)).strip()


@fitxes_bp.route('/fitxes/control-revisions', methods=['GET'])
@login_required
def control_revisions():
    """Retorna dades consolidades de control de revisions (com l'Excel PR09.02)."""
    try:
        dades = _build_control_revisions()
    except Exception as e:
        return jsonify({'error': f'Error construint dades: {str(e)}'}), 500

    # Estadístiques
    from datetime import timedelta
    ara = datetime.now(timezone.utc)
    limit_caducitat = ara - timedelta(days=730)  # 2 anys

    total = len(dades)
    publicades = sum(1 for d in dades if d['estat'] == 'publicada')
    esborranys = sum(1 for d in dades if d['estat'] == 'esborrany')
    obsoletes = sum(1 for d in dades if d['estat'] == 'obsoleta')
    en_revisio = sum(1 for d in dades if d.get('estat_versio') == 'en_revisio')

    caducades = 0
    for d in dades:
        if d['data_comprovacio']:
            try:
                from datetime import datetime as dt_cls
                data_comp = dt_cls.strptime(d['data_comprovacio'], '%d/%m/%Y')
                if data_comp < limit_caducitat.replace(tzinfo=None):
                    caducades += 1
                    d['caducada'] = True
            except (ValueError, TypeError):
                pass

    return jsonify({
        'fitxes': dades,
        'stats': {
            'total': total,
            'publicades': publicades,
            'esborranys': esborranys,
            'obsoletes': obsoletes,
            'en_revisio': en_revisio,
            'caducades': caducades,
        },
    })


@fitxes_bp.route('/fitxes/control-revisions/export', methods=['GET'])
@login_required
def exportar_control_revisions():
    """Exporta el control de revisions a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    dades = _build_control_revisions()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Control revisions'

    # Estils
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Títol
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'CONTROL DE REVISIONS DE FITXES TÈCNIQUES'
    ws['A1'].font = Font(bold=True, size=14, color='2F5496')
    ws['A2'] = f'Data exportació: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, color='666666')

    # Capçaleres
    headers = [
        'Codi', 'Nom article', 'Estat', 'Revisió', 'Data rev.',
        'Data comprov.', 'Client', 'Denominació jurídica', 'Composició',
        'Vida útil', 'Observacions', 'W', 'P/L', 'Proteïna', 'Gluten',
        'Cendres', 'Control de canvis',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    # Dades
    for row_idx, f in enumerate(dades, 5):
        # Historial de canvis formatat
        canvis = '\n'.join([
            f"Revisió {h['num_versio']} ({h['data']}): {h['descripcio']}"
            for h in f['historial']
        ])

        vals = [
            f['art_codi'], f['nom_producte'], f['estat'],
            f['revisio'], f['data_revisio'], f['data_comprovacio'],
            'Sí' if f['es_client'] else 'No',
            f['denominacio_juridica'], f['composicio'], f['vida_util'],
            f['observacions'], f['w'], f['pl'], f['proteina'],
            f['gluten'], f['cendres'], canvis,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Amplades columnes
    widths = [10, 35, 12, 8, 12, 12, 8, 25, 20, 20, 25, 12, 12, 12, 12, 12, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Control_revisions_{datetime.now().strftime("%Y%m%d")}.xlsx',
    )
