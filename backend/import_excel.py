"""
Script d'importació de l'Excel PR09.02 a la base de dades.
Importa el full 'Mestre productes elab actius'.

Ús:
    cd backend
    venv\Scripts\activate
    python import_excel.py
"""

import os
import re
import sys
from datetime import datetime, timezone

import openpyxl

# Afegir el directori backend al path
sys.path.insert(0, os.path.dirname(__file__))

from app import create_app, db
from app.models import FitxaTecnica, VersioFitxa

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), '..',
    'PR09.02 - Control revisions de fitxes tecniques de productes elaborats.xlsx'
)


def parse_control_canvis(text):
    """Parseja el text de control de canvis en revisions individuals.

    Format esperat:
        Revisió 0:
        Creació del producte
        Revisió 1:
        Canvis fets...
    """
    if not text:
        return []

    text = str(text).strip()
    # Separar per "Revisió N:" o "Revisió N :"
    parts = re.split(r'Revisi[oó]\s*(\d+)\s*:\s*', text, flags=re.IGNORECASE)

    revisions = []
    # parts[0] és text abans de la primera "Revisió N:", normalment buit
    # Després ve [num, text, num, text, ...]
    i = 1
    while i < len(parts) - 1:
        num = int(parts[i])
        desc = parts[i + 1].strip()
        if desc:
            revisions.append({'num': num, 'descripcio': desc})
        i += 2

    return revisions


def parse_es_client(val):
    """Determina si és fitxa de client a partir del camp Client."""
    if not val:
        return False, ''
    s = str(val).strip().lower()
    if s in ('no', 'na', ''):
        return False, ''
    # Qualsevol valor amb 'si' o un nom de client
    return True, str(val).strip()


def normalitzar_codi(codi):
    """Normalitza el codi d'article a string."""
    if codi is None:
        return None
    codi = str(codi).strip()
    if not codi:
        return None
    return codi


def importar():
    app = create_app()

    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: No es troba l'Excel a {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Mestre productes elab actius']

    with app.app_context():
        importats = 0
        errors = 0
        omesos = 0

        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=False):
            vals = {c.column: c.value for c in row if c.value is not None}

            # Columna A: Codi
            codi = normalitzar_codi(vals.get(1))
            if not codi:
                continue

            # Columna B: Nom Article
            nom = str(vals.get(2, '')).strip()
            if not nom:
                nom = f'Producte {codi}'

            # Columna C: Revisió
            rev_val = vals.get(3)
            if rev_val == 'NA' or rev_val is None:
                num_rev = 0
            else:
                try:
                    num_rev = int(rev_val)
                except (ValueError, TypeError):
                    num_rev = 0

            # Columna D: Data última Rev
            data_rev = vals.get(4)
            if isinstance(data_rev, datetime):
                data_rev_dt = data_rev.replace(tzinfo=timezone.utc)
            else:
                data_rev_dt = None

            # Columna E: Data Comprovació
            data_comp = vals.get(5)
            if isinstance(data_comp, datetime):
                data_comp_dt = data_comp.replace(tzinfo=timezone.utc)
            else:
                data_comp_dt = None

            # Columna F: Client
            es_client, client_info = parse_es_client(vals.get(6))

            # Columna G: Denominació jurídica
            denom_juridica = str(vals.get(7, '')).strip()

            # Columna H: Composició
            composicio = str(vals.get(8, '')).strip()

            # Columna I: Vida útil
            vida_util = str(vals.get(9, '')).strip()

            # Columna J: Observacions
            observacions = str(vals.get(10, '')).strip()

            # Columnes K-Q: Paràmetres tècnics
            w_min = vals.get(11, '')
            w_max = vals.get(12, '')
            pl_min = vals.get(13, '')
            pl_max = vals.get(14, '')
            proteina = vals.get(15, '')
            gluten = vals.get(16, '')
            cendres = vals.get(17, '')

            # Columna R: Control de canvis
            control_canvis = str(vals.get(18, '')).strip()

            # Comprovar si ja existeix
            existent = FitxaTecnica.query.filter_by(art_codi=codi).first()
            if existent:
                omesos += 1
                continue

            # Determinar estat
            obs_lower = observacions.lower()
            if 'inactiu' in obs_lower:
                estat = 'obsoleta'
            elif num_rev > 0:
                estat = 'publicada'
            else:
                estat = 'esborrany'

            try:
                # Crear fitxa
                fitxa = FitxaTecnica(
                    art_codi=codi,
                    nom_producte=nom,
                    categoria='producte elaborat',
                    estat=estat,
                    es_client=es_client,
                    observacions=observacions,
                    created_at=data_rev_dt or datetime.now(timezone.utc),
                    updated_at=data_rev_dt or datetime.now(timezone.utc),
                    created_by='importació Excel',
                )
                db.session.add(fitxa)
                db.session.flush()

                # Construir contingut JSON de la versió
                contingut = {}
                if denom_juridica:
                    contingut['denominacio_juridica'] = denom_juridica
                if composicio:
                    contingut['ingredients'] = composicio
                if vida_util:
                    contingut['vida_util'] = vida_util
                if client_info:
                    contingut['client_info'] = client_info

                # Paràmetres reològics
                reologiques = []
                if w_min or w_max:
                    w_str = f"{w_min}" if w_min == w_max or not w_max else f"{w_min}-{w_max}"
                    reologiques.append({
                        'parametre': 'W',
                        'valor': str(w_str).replace('None', ''),
                    })
                if pl_min or pl_max:
                    pl_str = f"{pl_min}" if pl_min == pl_max or not pl_max else f"{pl_min}-{pl_max}"
                    reologiques.append({
                        'parametre': 'P/L',
                        'valor': str(pl_str).replace('None', ''),
                    })
                if reologiques:
                    contingut['reologiques'] = reologiques

                # Paràmetres fisicoquímics
                fisicoquimiques = []
                if proteina:
                    fisicoquimiques.append({
                        'parametre': 'Proteïna',
                        'valor': str(proteina),
                    })
                if gluten:
                    fisicoquimiques.append({
                        'parametre': 'Gluten',
                        'valor': str(gluten),
                    })
                if cendres:
                    fisicoquimiques.append({
                        'parametre': 'Cendres',
                        'valor': str(cendres),
                    })
                if fisicoquimiques:
                    contingut['fisicoquimiques'] = fisicoquimiques

                # Parsejar historial de canvis per crear versions
                revisions = parse_control_canvis(control_canvis)

                if revisions:
                    # Crear una versió per cada revisió documentada
                    for rev_info in revisions:
                        versio = VersioFitxa(
                            fitxa_id=fitxa.id,
                            num_versio=rev_info['num'],
                            descripcio_canvi=rev_info['descripcio'],
                            contingut=contingut if rev_info['num'] == num_rev else None,
                            data_comprovacio=data_comp_dt if rev_info['num'] == num_rev else None,
                            created_at=data_rev_dt if rev_info['num'] == num_rev else None,
                            created_by='importació Excel',
                            activa=(rev_info['num'] == num_rev),
                        )
                        db.session.add(versio)
                else:
                    # Si no hi ha historial, crear una versió única
                    versio = VersioFitxa(
                        fitxa_id=fitxa.id,
                        num_versio=max(num_rev, 0),
                        descripcio_canvi='Importat des d\'Excel (sense historial)',
                        contingut=contingut,
                        data_comprovacio=data_comp_dt,
                        created_at=data_rev_dt or datetime.now(timezone.utc),
                        created_by='importació Excel',
                        activa=True,
                    )
                    db.session.add(versio)

                db.session.commit()
                importats += 1

            except Exception as e:
                db.session.rollback()
                errors += 1
                print(f"  ERROR {codi} ({nom}): {e}")

        print(f"\n{'='*50}")
        print(f"Importació completada!")
        print(f"  Importats: {importats}")
        print(f"  Omesos (ja existien): {omesos}")
        print(f"  Errors: {errors}")
        print(f"  Total processats: {importats + omesos + errors}")


if __name__ == '__main__':
    importar()
